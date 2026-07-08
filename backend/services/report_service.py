from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
import os

from models import (Order, OrderItem, Payment, Product, Category, User, Shift,
                    Customer, Inventory)


class ReportService:
    def __init__(self, db: Session):
        self.db = db

    # ── SAVDO ─────────────────────────────────────────────────────────────────────
    def generate_sales_report(self, date_from: datetime, date_to: datetime,
                               branch_id: Optional[int] = None,
                               current_user=None) -> Dict[str, Any]:
        from deps import apply_tenant_filter

        q = self.db.query(Payment).filter(
            Payment.created_at >= date_from,
            Payment.created_at <= date_to,
            Payment.status == "paid"
        )
        if current_user:
            q = apply_tenant_filter(q, Payment, current_user)
        if branch_id:
            q = q.filter(Payment.branch_id == branch_id)

        total = q.with_entities(func.sum(Payment.amount)).scalar() or 0
        count = q.with_entities(func.count(Payment.id)).scalar() or 0

        daily = q.with_entities(
            func.date(Payment.created_at).label('day'),
            func.sum(Payment.amount).label('total_val'),
            func.count(Payment.id).label('cnt')
        ).group_by(func.date(Payment.created_at)).order_by('day').all()

        methods = q.with_entities(
            Payment.method,
            func.sum(Payment.amount).label('total_val'),
            func.count(Payment.id).label('cnt')
        ).group_by(Payment.method).all()

        oq = self.db.query(Order).filter(
            Order.created_at >= date_from,
            Order.created_at <= date_to,
            Order.status == "completed"
        )
        if current_user:
            oq = apply_tenant_filter(oq, Order, current_user)
        if branch_id:
            oq = oq.filter(Order.branch_id == branch_id)
        orders_count = oq.count()

        return {
            "date_from": date_from.strftime("%Y-%m-%d"),
            "date_to": date_to.strftime("%Y-%m-%d"),
            "total_sales": float(total),
            "transactions_count": int(count),
            "orders_count": orders_count,
            "avg_check": float(total) / orders_count if orders_count else 0,
            "daily_sales": [{"date": str(r.day), "total": float(r.total_val), "count": r.cnt} for r in daily],
            "payment_methods": [{"method": m.method, "total": float(m.total_val), "count": m.cnt} for m in methods],
        }

    # ── FOYDA ─────────────────────────────────────────────────────────────────────
    def generate_profit_report(self, date_from: datetime, date_to: datetime,
                                branch_id: Optional[int] = None,
                                current_user=None) -> Dict[str, Any]:
        from deps import apply_tenant_filter

        q = self.db.query(
            func.date(Order.created_at).label('d'),
            func.sum(OrderItem.total_price).label('revenue'),
            func.sum(OrderItem.quantity * OrderItem.unit_cost).label('cost'),
            func.count(func.distinct(Order.id)).label('orders')
        ).join(Order, Order.id == OrderItem.order_id).filter(
            Order.created_at >= date_from,
            Order.created_at <= date_to,
            Order.status == "completed"
        )
        if current_user:
            q = apply_tenant_filter(q, Order, current_user)
        if branch_id:
            q = q.filter(Order.branch_id == branch_id)
        daily = q.group_by(func.date(Order.created_at)).order_by('d').all()

        total_revenue = sum(float(d.revenue or 0) for d in daily)
        total_cost = sum(float(d.cost or 0) for d in daily)
        total_profit = total_revenue - total_cost
        margin = (total_profit / total_revenue * 100) if total_revenue else 0

        pq = self.db.query(
            Product.name,
            func.sum(OrderItem.total_price).label('revenue'),
            func.sum(OrderItem.quantity * OrderItem.unit_cost).label('cost'),
            func.sum(OrderItem.quantity).label('qty')
        ).join(OrderItem, OrderItem.product_id == Product.id
        ).join(Order, Order.id == OrderItem.order_id).filter(
            Order.created_at >= date_from,
            Order.created_at <= date_to,
            Order.status == "completed"
        )
        if current_user:
            pq = apply_tenant_filter(pq, Order, current_user)
        if branch_id:
            pq = pq.filter(Order.branch_id == branch_id)
        products = pq.group_by(Product.id, Product.name).order_by(
            (func.sum(OrderItem.total_price) - func.sum(OrderItem.quantity * OrderItem.unit_cost)).desc()
        ).limit(20).all()

        return {
            "date_from": date_from.strftime("%Y-%m-%d"),
            "date_to": date_to.strftime("%Y-%m-%d"),
            "total_revenue": total_revenue,
            "total_cost": total_cost,
            "total_profit": total_profit,
            "margin_pct": round(margin, 1),
            "daily": [
                {
                    "date": str(d.d),
                    "revenue": float(d.revenue or 0),
                    "cost": float(d.cost or 0),
                    "profit": float(d.revenue or 0) - float(d.cost or 0),
                    "orders": d.orders
                }
                for d in daily
            ],
            "top_products": [
                {
                    "name": p.name,
                    "revenue": float(p.revenue or 0),
                    "cost": float(p.cost or 0),
                    "profit": float(p.revenue or 0) - float(p.cost or 0),
                    "qty": float(p.qty or 0),
                    "margin_pct": round((float(p.revenue or 0) - float(p.cost or 0)) / max(float(p.revenue or 1), 1) * 100, 1)
                }
                for p in products
            ]
        }

    # ── OMBOR ─────────────────────────────────────────────────────────────────────
    def generate_inventory_report(self, branch_id: Optional[int] = None,
                                   current_user=None) -> Dict[str, Any]:
        from deps import apply_tenant_filter

        q = self.db.query(Inventory)
        if current_user:
            q = apply_tenant_filter(q, Inventory, current_user)
        if branch_id:
            q = q.filter(Inventory.branch_id == branch_id)
        items = q.all()

        total_value = sum((i.quantity or 0) * (i.product.cost_price if i.product else 0) for i in items)
        low_stock = [i for i in items if 0 < (i.quantity or 0) < (i.min_threshold or 5)]
        out_of_stock = [i for i in items if (i.quantity or 0) <= 0]

        def _item(i):
            return {
                "id": i.id,
                "product_id": i.product_id,
                "product_name": i.product.name if i.product else "—",
                "quantity": float(i.quantity or 0),
                "unit": i.unit or "dona",
                "min_threshold": float(i.min_threshold or 0),
                "value": float((i.quantity or 0) * (i.product.cost_price if i.product else 0)),
                "is_low": 0 < (i.quantity or 0) < (i.min_threshold or 5),
                "is_out": (i.quantity or 0) <= 0
            }

        return {
            "total_items": len(items),
            "total_value": total_value,
            "low_stock_count": len(low_stock),
            "out_of_stock_count": len(out_of_stock),
            "items": [_item(i) for i in sorted(items, key=lambda x: x.product.name if x.product else '')],
            "low_stock": [_item(i) for i in low_stock],
        }

    # ── MIJOZLAR ──────────────────────────────────────────────────────────────────
    def generate_customers_report(self, date_from: datetime, date_to: datetime,
                                   branch_id: Optional[int] = None,
                                   current_user=None) -> Dict[str, Any]:
        from deps import apply_tenant_filter

        cq = self.db.query(Customer).filter(
            Customer.created_at >= date_from,
            Customer.created_at <= date_to
        )
        if current_user:
            cq = apply_tenant_filter(cq, Customer, current_user)
        new_count = cq.count()

        tq = self.db.query(Customer).filter(Customer.total_spent > 0)
        if current_user:
            tq = apply_tenant_filter(tq, Customer, current_user)
        top = tq.order_by(Customer.total_spent.desc()).limit(20).all()

        oq = self.db.query(
            Order.customer_id,
            func.count(Order.id).label('orders_count'),
            func.sum(Order.final_amount).label('spent')
        ).filter(
            Order.created_at >= date_from,
            Order.created_at <= date_to,
            Order.status == "completed",
            Order.customer_id != None
        )
        if current_user:
            oq = apply_tenant_filter(oq, Order, current_user)
        if branch_id:
            oq = oq.filter(Order.branch_id == branch_id)
        period_orders = oq.group_by(Order.customer_id).all()

        loyal_q = self.db.query(func.count(Customer.id)).filter(Customer.total_visits >= 3)
        if current_user:
            loyal_q = apply_tenant_filter(loyal_q, Customer, current_user)
        total_loyal = loyal_q.scalar() or 0

        return {
            "date_from": date_from.strftime("%Y-%m-%d"),
            "date_to": date_to.strftime("%Y-%m-%d"),
            "new_customers": new_count,
            "total_loyal": int(total_loyal),
            "active_in_period": len(period_orders),
            "top_customers": [
                {
                    "id": c.id,
                    "name": c.name,
                    "phone": c.phone or "",
                    "total_visits": c.total_visits or 0,
                    "total_spent": float(c.total_spent or 0),
                    "points": c.points or 0
                }
                for c in top
            ],
        }

    # ── SOLIQ ─────────────────────────────────────────────────────────────────────
    def generate_tax_report(self, date_from: datetime, date_to: datetime,
                             branch_id: Optional[int] = None,
                             current_user=None) -> Dict[str, Any]:
        from deps import apply_tenant_filter

        q = self.db.query(Payment).filter(
            Payment.created_at >= date_from,
            Payment.created_at <= date_to,
            Payment.status == "paid"
        )
        if current_user:
            q = apply_tenant_filter(q, Payment, current_user)
        if branch_id:
            q = q.filter(Payment.branch_id == branch_id)

        total = float(q.with_entities(func.sum(Payment.amount)).scalar() or 0)
        count = int(q.with_entities(func.count(Payment.id)).scalar() or 0)

        daily = q.with_entities(
            func.date(Payment.created_at).label('day'),
            func.sum(Payment.amount).label('total_val'),
            func.count(Payment.id).label('cnt')
        ).group_by(func.date(Payment.created_at)).order_by('day').all()

        methods = q.with_entities(
            Payment.method,
            func.sum(Payment.amount).label('total_val'),
            func.count(Payment.id).label('cnt')
        ).group_by(Payment.method).all()

        vat_rate = 0.12
        vat_amount = total * vat_rate / (1 + vat_rate)
        net_amount = total - vat_amount

        return {
            "date_from": date_from.strftime("%Y-%m-%d"),
            "date_to": date_to.strftime("%Y-%m-%d"),
            "gross_revenue": total,
            "transactions_count": count,
            "vat_rate_pct": 12,
            "vat_amount": round(vat_amount, 2),
            "net_amount": round(net_amount, 2),
            "daily": [{"date": str(r.day), "gross": float(r.total_val), "count": r.cnt} for r in daily],
            "by_method": [{"method": m.method, "total": float(m.total_val), "count": m.cnt} for m in methods],
        }

    # ── MAHSULOTLAR ───────────────────────────────────────────────────────────────
    def generate_products_report(self, date_from: datetime, date_to: datetime,
                                  limit: int = 50, branch_id: Optional[int] = None,
                                  current_user=None) -> Dict[str, Any]:
        q = self.db.query(
            Product.id, Product.name,
            Category.name.label('category'),
            func.sum(OrderItem.quantity).label('quantity'),
            func.sum(OrderItem.total_price).label('revenue'),
            func.count(func.distinct(Order.id)).label('orders_count')
        ).join(Category, Category.id == Product.category_id
        ).join(OrderItem, OrderItem.product_id == Product.id
        ).join(Order, Order.id == OrderItem.order_id).filter(
            Order.created_at >= date_from,
            Order.created_at <= date_to,
            Order.status == "completed"
        )
        if current_user:
            from deps import apply_tenant_filter
            q = apply_tenant_filter(q, Order, current_user)
        if branch_id:
            q = q.filter(Order.branch_id == branch_id)

        products = q.group_by(Product.id, Product.name, Category.name
        ).order_by(func.sum(OrderItem.quantity).desc()).limit(limit).all()

        return {
            "date_from": date_from.strftime("%Y-%m-%d"),
            "date_to": date_to.strftime("%Y-%m-%d"),
            "products": [
                {
                    "id": p.id, "name": p.name, "category": p.category,
                    "quantity": int(p.quantity or 0),
                    "revenue": float(p.revenue or 0),
                    "orders_count": p.orders_count
                }
                for p in products
            ]
        }

    # ── XODIMLAR ──────────────────────────────────────────────────────────────────
    def generate_staff_report(self, date_from: datetime, date_to: datetime,
                               branch_id: Optional[int] = None,
                               current_user=None) -> Dict[str, Any]:
        q = self.db.query(
            User.id, User.full_name,
            func.count(Order.id).label('orders_count'),
            func.sum(Order.final_amount).label('total_sales'),
            func.avg(Order.final_amount).label('avg_check')
        ).join(Order, Order.waiter_id == User.id).filter(
            Order.created_at >= date_from,
            Order.created_at <= date_to,
            Order.status == "completed"
        )
        if current_user:
            from deps import apply_tenant_filter
            q = apply_tenant_filter(q, Order, current_user)
        if branch_id:
            q = q.filter(Order.branch_id == branch_id)

        staff = q.group_by(User.id, User.full_name
        ).order_by(func.sum(Order.final_amount).desc()).all()

        return {
            "date_from": date_from.strftime("%Y-%m-%d"),
            "date_to": date_to.strftime("%Y-%m-%d"),
            "staff": [
                {
                    "id": s.id, "name": s.full_name,
                    "orders_count": s.orders_count,
                    "total_sales": float(s.total_sales or 0),
                    "avg_check": float(s.avg_check or 0)
                }
                for s in staff
            ]
        }

    # ── SMENA ─────────────────────────────────────────────────────────────────────
    def generate_shift_report(self, shift_id: Optional[int] = None,
                               user_id: Optional[int] = None,
                               date: Optional[str] = None,
                               current_user=None) -> Dict[str, Any]:
        query = self.db.query(Shift)
        if current_user:
            from deps import apply_tenant_filter
            query = apply_tenant_filter(query, Shift, current_user)

        if shift_id:
            query = query.filter(Shift.id == shift_id)
        elif user_id and date:
            target_date = datetime.strptime(date, "%Y-%m-%d")
            query = query.filter(
                Shift.user_id == user_id,
                func.date(Shift.start_time) == target_date.date()
            )

        shifts = query.all()
        result = []
        for shift in shifts:
            orders = self.db.query(Order).filter(
                Order.waiter_id == shift.user_id,
                Order.created_at >= shift.start_time,
                Order.created_at <= (shift.end_time or datetime.now())
            ).all()
            total_sales = sum(o.final_amount for o in orders if o.status == "completed")
            result.append({
                "id": shift.id,
                "user_name": shift.user.full_name if shift.user else None,
                "start_time": shift.start_time.isoformat() if shift.start_time else None,
                "end_time": shift.end_time.isoformat() if shift.end_time else None,
                "starting_cash": shift.starting_cash,
                "ending_cash": shift.ending_cash,
                "total_sales": total_sales,
                "orders_count": len(orders),
                "cash_sales": shift.cash_sales,
                "card_sales": shift.card_sales
            })
        return {"shifts": result}

    # ── EXCEL EXPORT ──────────────────────────────────────────────────────────────
    def export_excel(self, report_type: str, date_from: datetime, date_to: datetime,
                     branch_id: Optional[int] = None, current_user=None) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active

        def hdr(col, row, val):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = PatternFill("solid", fgColor="1a1a2e")
            c.font = Font(bold=True, color="D4AF37", size=10)
            c.alignment = Alignment(horizontal="center")

        def mfmt(cell):
            cell.number_format = '#,##0'

        title_font = Font(bold=True, size=14, color="1a1a2e")
        gold_font = Font(bold=True, color="D4AF37", size=11)

        if report_type == "sales":
            data = self.generate_sales_report(date_from, date_to, branch_id, current_user)
            ws.title = "Savdo"
            ws['A1'] = f"SAVDO HISOBOTI  {data['date_from']} — {data['date_to']}"
            ws['A1'].font = title_font
            ws.merge_cells('A1:E1')
            ws['A3'] = "Jami savdo"; ws['B3'] = data['total_sales']; mfmt(ws['B3'])
            ws['A4'] = "Buyurtmalar"; ws['B4'] = data['orders_count']
            ws['A5'] = "O'rtacha chek"; ws['B5'] = round(data['avg_check']); mfmt(ws['B5'])
            ws['A7'] = "Kunlik savdo"; ws['A7'].font = gold_font
            for ci, h in enumerate(["Sana", "Jami (so'm)", "Tranzaksiyalar"], 1): hdr(ci, 8, h)
            for ri, d in enumerate(data['daily_sales'], 9):
                ws.cell(ri, 1, d['date']); mfmt(ws.cell(ri, 2, d['total'])); ws.cell(ri, 3, d['count'])
            r = 9 + len(data['daily_sales']) + 1
            ws.cell(r, 1, "To'lov usullari").font = gold_font
            for ci, h in enumerate(["Usul", "Jami", "Soni"], 1): hdr(ci, r+1, h)
            for ri2, m in enumerate(data['payment_methods'], r+2):
                ws.cell(ri2, 1, m['method']); mfmt(ws.cell(ri2, 2, m['total'])); ws.cell(ri2, 3, m['count'])

        elif report_type == "profit":
            data = self.generate_profit_report(date_from, date_to, branch_id, current_user)
            ws.title = "Foyda"
            ws['A1'] = f"FOYDA HISOBOTI  {data['date_from']} — {data['date_to']}"
            ws['A1'].font = title_font; ws.merge_cells('A1:F1')
            ws['A3'] = "Daromad"; ws['B3'] = data['total_revenue']; mfmt(ws['B3'])
            ws['A4'] = "Xarajat";  ws['B4'] = data['total_cost'];    mfmt(ws['B4'])
            ws['A5'] = "Foyda";    ws['B5'] = data['total_profit'];   mfmt(ws['B5'])
            ws['A6'] = "Marja";    ws['B6'] = f"{data['margin_pct']}%"
            ws['A8'] = "Kunlik"; ws['A8'].font = gold_font
            for ci, h in enumerate(["Sana", "Daromad", "Xarajat", "Foyda", "Buyurtmalar"], 1): hdr(ci, 9, h)
            for ri, d in enumerate(data['daily'], 10):
                ws.cell(ri, 1, d['date']); mfmt(ws.cell(ri, 2, d['revenue']))
                mfmt(ws.cell(ri, 3, d['cost'])); mfmt(ws.cell(ri, 4, d['profit'])); ws.cell(ri, 5, d['orders'])
            r = 10 + len(data['daily']) + 1
            ws.cell(r, 1, "Top mahsulotlar").font = gold_font
            for ci, h in enumerate(["Mahsulot", "Daromad", "Xarajat", "Foyda", "Soni", "Marja%"], 1): hdr(ci, r+1, h)
            for ri2, p in enumerate(data['top_products'], r+2):
                ws.cell(ri2, 1, p['name']); mfmt(ws.cell(ri2, 2, p['revenue']))
                mfmt(ws.cell(ri2, 3, p['cost'])); mfmt(ws.cell(ri2, 4, p['profit']))
                ws.cell(ri2, 5, p['qty']); ws.cell(ri2, 6, f"{p['margin_pct']}%")

        elif report_type == "inventory":
            data = self.generate_inventory_report(branch_id, current_user)
            ws.title = "Ombor"
            ws['A1'] = "OMBOR HISOBOTI"; ws['A1'].font = title_font
            ws['A3'] = "Jami pozitsiyalar"; ws['B3'] = data['total_items']
            ws['A4'] = "Jami qiymat"; ws['B4'] = data['total_value']; mfmt(ws['B4'])
            ws['A5'] = "Kam qoldiq"; ws['B5'] = data['low_stock_count']
            ws['A6'] = "Tugagan"; ws['B6'] = data['out_of_stock_count']
            ws['A8'] = "Barchasi"; ws['A8'].font = gold_font
            for ci, h in enumerate(["Mahsulot", "Miqdor", "O'lchov", "Min", "Qiymat", "Holat"], 1): hdr(ci, 9, h)
            for ri, i in enumerate(data['items'], 10):
                ws.cell(ri, 1, i['product_name']); ws.cell(ri, 2, i['quantity'])
                ws.cell(ri, 3, i['unit']); ws.cell(ri, 4, i['min_threshold'])
                mfmt(ws.cell(ri, 5, i['value'])); ws.cell(ri, 6, "KAM" if i['is_low'] else ("TUG" if i['is_out'] else "OK"))

        elif report_type == "customers":
            data = self.generate_customers_report(date_from, date_to, branch_id, current_user)
            ws.title = "Mijozlar"
            ws['A1'] = f"MIJOZLAR HISOBOTI  {data['date_from']} — {data['date_to']}"
            ws['A1'].font = title_font; ws.merge_cells('A1:E1')
            ws['A3'] = "Yangi mijozlar"; ws['B3'] = data['new_customers']
            ws['A4'] = "Faol (davr)"; ws['B4'] = data['active_in_period']
            ws['A5'] = "Sodiq (3+ tashrif)"; ws['B5'] = data['total_loyal']
            ws['A7'] = "Top mijozlar"; ws['A7'].font = gold_font
            for ci, h in enumerate(["Ism", "Telefon", "Tashriflar", "Sarflagan", "Ball"], 1): hdr(ci, 8, h)
            for ri, c in enumerate(data['top_customers'], 9):
                ws.cell(ri, 1, c['name']); ws.cell(ri, 2, c['phone'])
                ws.cell(ri, 3, c['total_visits']); mfmt(ws.cell(ri, 4, c['total_spent'])); ws.cell(ri, 5, c['points'])

        elif report_type == "staff":
            data = self.generate_staff_report(date_from, date_to, branch_id, current_user)
            ws.title = "Xodimlar"
            ws['A1'] = f"XODIMLAR HISOBOTI  {data['date_from']} — {data['date_to']}"
            ws['A1'].font = title_font; ws.merge_cells('A1:D1')
            for ci, h in enumerate(["Xodim", "Buyurtmalar", "Jami savdo", "O'rtacha chek"], 1): hdr(ci, 3, h)
            for ri, s in enumerate(data['staff'], 4):
                ws.cell(ri, 1, s['name']); ws.cell(ri, 2, s['orders_count'])
                mfmt(ws.cell(ri, 3, s['total_sales'])); mfmt(ws.cell(ri, 4, s['avg_check']))

        elif report_type == "tax":
            data = self.generate_tax_report(date_from, date_to, branch_id, current_user)
            ws.title = "Soliq"
            ws['A1'] = f"SOLIQ HISOBOTI  {data['date_from']} — {data['date_to']}"
            ws['A1'].font = title_font; ws.merge_cells('A1:C1')
            ws['A3'] = "Brutto daromad"; ws['B3'] = data['gross_revenue']; mfmt(ws['B3'])
            ws['A4'] = f"QQS ({data['vat_rate_pct']}%)"; ws['B4'] = data['vat_amount']; mfmt(ws['B4'])
            ws['A5'] = "Netto"; ws['B5'] = data['net_amount']; mfmt(ws['B5'])
            ws['A6'] = "Tranzaksiyalar"; ws['B6'] = data['transactions_count']
            ws['A8'] = "Kunlik"; ws['A8'].font = gold_font
            for ci, h in enumerate(["Sana", "Brutto", "Tranzaksiyalar"], 1): hdr(ci, 9, h)
            for ri, d in enumerate(data['daily'], 10):
                ws.cell(ri, 1, d['date']); mfmt(ws.cell(ri, 2, d['gross'])); ws.cell(ri, 3, d['count'])

        # Auto column width
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            w = max((len(str(c.value)) + 4 for c in col if c.value), default=12)
            ws.column_dimensions[letter].width = min(w, 42)

        fname = f"{report_type}_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
        fpath = os.path.join("static", "exports", fname)
        os.makedirs(os.path.dirname(fpath), exist_ok=True)
        wb.save(fpath)
        return f"/static/exports/{fname}"

    # ── PDF EXPORT ────────────────────────────────────────────────────────────────
    def export_pdf(self, report_type: str, date_from: datetime, date_to: datetime,
                   branch_id: Optional[int] = None, current_user=None) -> str:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        GOLD = colors.Color(0.831, 0.686, 0.216)
        DARK = colors.Color(0.102, 0.102, 0.180)

        fname = f"{report_type}_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.pdf"
        fpath = os.path.join("static", "exports", fname)
        os.makedirs(os.path.dirname(fpath), exist_ok=True)

        doc = SimpleDocTemplate(fpath, pagesize=A4,
                                 leftMargin=1.5*cm, rightMargin=1.5*cm,
                                 topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('T', parent=styles['Title'], textColor=DARK, fontSize=15, spaceAfter=10)
        sub_style = ParagraphStyle('S', parent=styles['Normal'],
                                    textColor=GOLD, fontSize=11, spaceAfter=5, fontName='Helvetica-Bold')
        normal = styles['Normal']

        tbl_hdr = [
            ('BACKGROUND', (0, 0), (-1, 0), DARK),
            ('TEXTCOLOR', (0, 0), (-1, 0), GOLD),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.96, 0.96, 0.96)]),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.Color(0.85, 0.85, 0.85)),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]

        def fmt(n):
            try: return f"{int(n):,}".replace(',', ' ')
            except: return str(n)

        def tbl(rows, widths):
            t = Table(rows, colWidths=widths)
            t.setStyle(TableStyle(tbl_hdr))
            return t

        elements = []

        if report_type == "sales":
            data = self.generate_sales_report(date_from, date_to, branch_id, current_user)
            elements += [
                Paragraph("SAVDO HISOBOTI", title_style),
                Paragraph(f"{data['date_from']} — {data['date_to']}", normal),
                Spacer(1, 0.3*cm),
                tbl([["Ko'rsatkich", "Qiymat"],
                     ["Jami savdo", fmt(data['total_sales']) + " so'm"],
                     ["Buyurtmalar", str(data['orders_count'])],
                     ["O'rtacha chek", fmt(data['avg_check']) + " so'm"]],
                    [9*cm, 8*cm]),
                Spacer(1, 0.4*cm),
                Paragraph("Kunlik savdo", sub_style),
                tbl([["Sana", "Jami (so'm)", "Tranzaksiyalar"]] +
                    [[d['date'], fmt(d['total']), str(d['count'])] for d in data['daily_sales']],
                    [5*cm, 8*cm, 5*cm]),
                Spacer(1, 0.4*cm),
                Paragraph("To'lov usullari", sub_style),
                tbl([["Usul", "Jami", "Soni"]] +
                    [[m['method'], fmt(m['total']), str(m['count'])] for m in data['payment_methods']],
                    [6*cm, 7*cm, 5*cm]),
            ]

        elif report_type == "profit":
            data = self.generate_profit_report(date_from, date_to, branch_id, current_user)
            elements += [
                Paragraph("FOYDA HISOBOTI", title_style),
                Paragraph(f"{data['date_from']} — {data['date_to']}", normal),
                Spacer(1, 0.3*cm),
                tbl([["Ko'rsatkich", "Qiymat"],
                     ["Daromad", fmt(data['total_revenue']) + " so'm"],
                     ["Xarajat", fmt(data['total_cost']) + " so'm"],
                     ["Foyda", fmt(data['total_profit']) + " so'm"],
                     ["Marja", f"{data['margin_pct']}%"]],
                    [9*cm, 8*cm]),
                Spacer(1, 0.4*cm),
                Paragraph("Top foydali mahsulotlar", sub_style),
                tbl([["Mahsulot", "Daromad", "Xarajat", "Foyda", "Marja%"]] +
                    [[p['name'], fmt(p['revenue']), fmt(p['cost']), fmt(p['profit']), f"{p['margin_pct']}%"]
                     for p in data['top_products'][:15]],
                    [6*cm, 3.5*cm, 3.5*cm, 3.5*cm, 2.5*cm]),
            ]

        elif report_type == "inventory":
            data = self.generate_inventory_report(branch_id, current_user)
            elements += [
                Paragraph("OMBOR HISOBOTI", title_style),
                Spacer(1, 0.3*cm),
                tbl([["Ko'rsatkich", "Qiymat"],
                     ["Jami pozitsiyalar", str(data['total_items'])],
                     ["Jami qiymat", fmt(data['total_value']) + " so'm"],
                     ["Kam qoldiq", str(data['low_stock_count'])],
                     ["Tugagan", str(data['out_of_stock_count'])]],
                    [9*cm, 8*cm]),
                Spacer(1, 0.4*cm),
                Paragraph("Ombor holati", sub_style),
                tbl([["Mahsulot", "Miqdor", "Birlik", "Min", "Qiymat", "Holat"]] +
                    [[i['product_name'], str(i['quantity']), i['unit'], str(i['min_threshold']),
                      fmt(i['value']), "KAM" if i['is_low'] else ("TUG" if i['is_out'] else "OK")]
                     for i in data['items']],
                    [6*cm, 2.5*cm, 2.5*cm, 2*cm, 3.5*cm, 2.5*cm]),
            ]

        elif report_type == "customers":
            data = self.generate_customers_report(date_from, date_to, branch_id, current_user)
            elements += [
                Paragraph("MIJOZLAR HISOBOTI", title_style),
                Paragraph(f"{data['date_from']} — {data['date_to']}", normal),
                Spacer(1, 0.3*cm),
                tbl([["Ko'rsatkich", "Qiymat"],
                     ["Yangi mijozlar", str(data['new_customers'])],
                     ["Faol mijozlar (davr)", str(data['active_in_period'])],
                     ["Sodiq (3+ tashrif)", str(data['total_loyal'])]],
                    [9*cm, 8*cm]),
                Spacer(1, 0.4*cm),
                Paragraph("Top mijozlar", sub_style),
                tbl([["Ism", "Telefon", "Tashriflar", "Sarflagan", "Ball"]] +
                    [[c['name'], c['phone'], str(c['total_visits']), fmt(c['total_spent']), str(c['points'])]
                     for c in data['top_customers'][:20]],
                    [5*cm, 4*cm, 3*cm, 4*cm, 3*cm]),
            ]

        elif report_type == "staff":
            data = self.generate_staff_report(date_from, date_to, branch_id, current_user)
            elements += [
                Paragraph("XODIMLAR HISOBOTI", title_style),
                Paragraph(f"{data['date_from']} — {data['date_to']}", normal),
                Spacer(1, 0.4*cm),
                tbl([["Xodim", "Buyurtmalar", "Jami savdo", "O'rtacha chek"]] +
                    [[s['name'], str(s['orders_count']), fmt(s['total_sales']), fmt(s['avg_check'])]
                     for s in data['staff']],
                    [6*cm, 4*cm, 5*cm, 5*cm]),
            ]

        elif report_type == "tax":
            data = self.generate_tax_report(date_from, date_to, branch_id, current_user)
            elements += [
                Paragraph("SOLIQ / FISKAL HISOBOTI", title_style),
                Paragraph(f"{data['date_from']} — {data['date_to']}", normal),
                Spacer(1, 0.3*cm),
                tbl([["Ko'rsatkich", "Qiymat"],
                     ["Brutto daromad", fmt(data['gross_revenue']) + " so'm"],
                     [f"QQS ({data['vat_rate_pct']}%)", fmt(data['vat_amount']) + " so'm"],
                     ["Netto", fmt(data['net_amount']) + " so'm"],
                     ["Tranzaksiyalar", str(data['transactions_count'])]],
                    [9*cm, 8*cm]),
                Spacer(1, 0.4*cm),
                Paragraph("To'lov usullari bo'yicha", sub_style),
                tbl([["Usul", "Jami", "Soni"]] +
                    [[m['method'], fmt(m['total']), str(m['count'])] for m in data['by_method']],
                    [6*cm, 7*cm, 5*cm]),
                Spacer(1, 0.4*cm),
                Paragraph("Kunlik", sub_style),
                tbl([["Sana", "Brutto", "Tranzaksiyalar"]] +
                    [[d['date'], fmt(d['gross']), str(d['count'])] for d in data['daily']],
                    [5*cm, 8*cm, 5*cm]),
            ]

        doc.build(elements)
        return f"/static/exports/{fname}"

    # ── EXPORT DISPATCHER ─────────────────────────────────────────────────────────
    def export_report(self, report_type: str, date_from: datetime, date_to: datetime,
                      format: str, branch_id: Optional[int] = None,
                      current_user=None) -> str:
        if format == "excel":
            return self.export_excel(report_type, date_from, date_to, branch_id, current_user)
        elif format == "pdf":
            return self.export_pdf(report_type, date_from, date_to, branch_id, current_user)
        # csv fallback (legacy)
        import csv
        fname = f"{report_type}_{date_from.strftime('%Y%m%d')}.csv"
        fpath = os.path.join("static", "exports", fname)
        os.makedirs(os.path.dirname(fpath), exist_ok=True)
        data = self.generate_sales_report(date_from, date_to, branch_id, current_user)
        with open(fpath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            for k, v in data.items():
                writer.writerow([k, v])
        return f"/static/exports/{fname}"

    # ── DAILY (legacy compat) ─────────────────────────────────────────────────────
    def generate_daily_report(self, date: datetime, current_user=None) -> Dict[str, Any]:
        from deps import apply_tenant_filter
        start = date.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1)
        return self.generate_sales_report(start, end, current_user=current_user)
